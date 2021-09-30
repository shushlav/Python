# working_with_excel_files
#https://www.datacamp.com/community/tutorials/python-excel-tutorial

# Import pandas
import pandas as pd

# Assign spreadsheet filename to `file`
file = 'D:\Python\pandas\student_data2.xlsx'

# Load spreadsheet
xl = pd.ExcelFile(file)

# Print the sheet names
print(xl.sheet_names)       #['student_data2', 'new']

# Load a sheet into a DataFrame by name: df1
df1 = xl.parse('new', index_col=0)
print(df1)
                        # How To Write Pandas DataFrames to Excel Files
# Specify a writer
writer = pd.ExcelWriter('example.xlsx', engine='xlsxwriter')
#you add a sheet with the data to an existing workbook, which could have many worksheets in a workbook
# Write your DataFrame to a file
df1.to_excel(writer, 'Sheet1')
writer.save()

                    # A much better and a simple option is to write data in .csv extension
df1.to_csv("example.csv")

'''                    #  Using Conda Environment
# Install virtualenv
$ conda create --name excel python=3.5

# Activate `excel`
$ conda activate excel
# Install `openpyxl` in `excel`
$ pip install openpyxl
from openpyxl import load_workbook

# Go to the folder of your project
$ cd 'D:\Python\pandas\Project_ex'
# Load in the workbook
wb = load_workbook('student_data2.xlsx')
# Get sheet names
print(wb.sheetnames)

# Deactivate `excel`
$ conda deactivate                    
'''
from openpyxl import load_workbook     # a preffered way is PANDAS!!!!
wb = load_workbook('D:\Python\pandas\student_data2.xlsx')
# Get sheet names
print(wb.sheetnames)        #['student_data2', 'new', 'numbers']
# Get a sheet by name
sheet = wb['numbers']
# Print the sheet title
print('Sheet Title:',sheet.title)    # Sheet Title: numbers
# Get currently active sheet
anotherSheet = wb.active
# Check `anotherSheet`
print(anotherSheet)         # <Worksheet "numbers">
# Retrieve the value of a certain cell
print(sheet['A1'].value)    # M

c = sheet['B3']
print(c.value)              # 6
# Retrieve the row number of your element
print('Row No.:', c.row)    #Row No.: 3

# Retrieve the column number of your element
print('Column Letter:', c.column)    #Column Letter: 2   TODO -no letter, but number!!
print('Column Letter:', c.column_letter)    # B
# Retrieve the coordinates of the cell
print('Coordinates of cell:', c.coordinate)     #  Coordinates of cell: B3
print(sheet.cell(row=1, column=2).value)        # N
# Print a specific column
for i in range(1, 4):
     print(i, sheet.cell(row=i, column=2).value)
# Print row per row
for cellObj in sheet['A1':'C3']:
      for cell in cellObj:
              print(cell.coordinate, cell.value)
      print('--- END ---')     
# max_row and max_column.  =attributes general ways of making sure that you loaded in the data correctly
# Retrieve the maximum amount of rows
print('Max Rows:', sheet.max_row)       #Max Rows: 4

# Retrieve the maximum amount of columns
print('Max Columns:', sheet.max_column)    #Max Columns: 4